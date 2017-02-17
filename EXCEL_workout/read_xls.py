#! python3
# -*- coding: utf-8 -*-

#pip install openpyxl
# reading excel spreadsheets
import sys
import os
import openpyxl

workbook = openpyxl.load_workbook('Channels.xlsx')
print(type(workbook))
sheet = workbook.get_sheet_by_name('EN')
print(type(sheet))
sheetnames = workbook.get_sheet_names()
print(sheetnames)
cell = sheet['A1']
print(str(cell.value))
print(str(sheet['B2'].value))

for i in range(1, 10):
	print(i, sheet.cell(row=i, column=1).value)

http://192.168.1.4:554
 rtsp://192.168.1.4:554/ch0.h264

ifconfig ra0 192.168.1.100 
netmask 255.255.255.0
route add default gw 192.168.1.1